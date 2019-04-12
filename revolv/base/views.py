import json
import logging
from collections import OrderedDict

import mailchimp
import stripe
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.contrib.auth import views as auth_views
from django.contrib.auth.decorators import login_required
from django.core import serializers
from django.core.urlresolvers import reverse
from django.db.models import Q
from django.db.models import Sum, Count
from django.http import HttpResponse
from django.http import HttpResponseBadRequest
from django.http import HttpResponseRedirect
from django.shortcuts import redirect, render_to_response, get_object_or_404, render
from django.template.context import RequestContext
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.debug import sensitive_post_parameters
from django.views.decorators.http import require_http_methods
from django.views.generic import FormView, TemplateView, View
from revolv.base.forms import SignupForm, UpdateUser, RevolvUserProfileForm, SignInForm
from revolv.base.models import RevolvUserProfile
from revolv.base.users import UserDataMixin
from revolv.base.utils import ProjectGroup
from revolv.donor.views import humanize_integers, total_donations
from revolv.lib.mailer import send_revolv_email
from revolv.payments.models import Payment, Tip, RepaymentFragment, ReferralSourceTrack
from revolv.payments.models import UserReinvestment, AdminRepayment
from revolv.project.models import Category, Project, ProjectMatchingDonors, StripeDetails, AnonymousUserDonation
from revolv.project.utils import aggregate_stats
from revolv.solar_ed_week.models import HostEvent, BecomePartner
from revolv.tasks.sfdc import send_signup_info
from social.apps.django_app.default.models import UserSocialAuth

logger = logging.getLogger(__name__)
LIST_ID = settings.LIST_ID
NEWSLETTERS = "RE-Volv Newsletter"
ANNOUNCEMENTS = "RE-Volv Important Announcement"


class HomePageView(UserDataMixin, TemplateView):
    """
    Website home page.

    TODO: this view is deprecated - most of the context variables are not
    used anymore. Should be cleaned up.
    """
    template_name = 'base/home.html'
    FEATURED_PROJECT_TO_SHOW = 6

    def get_global_impacts(self):
        """
        Returns: A dictionary of RE-volv wide impact figures.
        """
        # carbon_saved_by_month = Project.objects.statistics().pounds_carbon_saved_per_month
        # Assume 20 year lifetime.
        # We use str() to avoid django adding commas to integer in the template.
        people_donated_sys_count = RevolvUserProfile.objects.exclude(project=None).count()
        people_donated_stat_Count = str(int(people_donated_sys_count))

        global_impacts = {
            # Users who have backed at least one project:
            'num_people_donated': people_donated_stat_Count,
            'num_projects': Project.objects.get_completed().count(),
            'num_people_affected':
                Project.objects.filter(project_status=Project.COMPLETED).aggregate(n=Sum('people_affected'))['n'],
            'co2_avoided': str(int(Project.objects.get_total_avoided_co2())),
        }
        return global_impacts

    def get_context_data(self, **kwargs):
        context = super(HomePageView, self).get_context_data(**kwargs)
        featured_projects = Project.objects.get_featured(HomePageView.FEATURED_PROJECT_TO_SHOW)
        active_projects = Project.objects.get_active()
        context["active_projects"] = filter(lambda p: p.amount_left > 0.0, active_projects)
        completed_projects = Project.objects.get_completed().reverse()
        context["first_project"] = active_projects[0] if len(active_projects) > 0 else None
        # Get top 6 featured projects, Changed to active Projects in final fix
        context["featured_projects"] = active_projects
        # accept return value from project/model.py and display it on project/home.html file
        context["completed_featured_projects"] = completed_projects
        context["active_projects_count"] = Project.objects.get_active().count()
        context["completed_projects_count"] = Project.objects.get_completed().count()
        context["total_donors_count"] = Payment.objects.total_distinct_organic_donors()
        context["global_impacts"] = self.get_global_impacts()
        return context


class DonationReportView(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/donation_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if request.user.revolvuserprofile.is_administrator() or request.user.revolvuserprofile.is_ambassador():
            return super(DonationReportView, self).dispatch(request, *args, **kwargs)
        return HttpResponseRedirect(reverse("dashboard"))

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(DonationReportView, self).get_context_data(**kwargs)
        return context


class DonationReportForProject(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/ambassador_donation_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_ambassador():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(DonationReportForProject, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(DonationReportForProject, self).get_context_data(**kwargs)
        return context


class RepaymentReport(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/repayment_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(RepaymentReport, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(RepaymentReport, self).get_context_data(**kwargs)
        return context


class BaseStaffDashboardView(UserDataMixin, TemplateView):
    """
    Base view for the administrator and ambassador dashboard views. The
    specific views in administrator/views.py and ambassador/views.py
    will inherit from this view.
    """

    def get_filter_args(self):
        """
        Return an array of arguments to pass to Project.objects.get_[drafted|proposed|active|completed].
        """
        return []

    def get_context_data(self, **kwargs):
        context = super(BaseStaffDashboardView, self).get_context_data(**kwargs)

        project_dict = OrderedDict()
        project_dict[ProjectGroup('Proposed Projects', "proposed")] = Project.objects.get_proposed(
            *self.get_filter_args())
        project_dict[ProjectGroup('Staged projects', "staged")] = Project.objects.get_staged(*self.get_filter_args())
        project_dict[ProjectGroup('Active Projects', "active")] = Project.objects.get_active(*self.get_filter_args())
        project_dict[ProjectGroup('Completed Projects', "completed")] = Project.objects.get_completed(
            *self.get_filter_args())

        context["project_dict"] = project_dict
        context["role"] = self.role or "donor"

        context['donated_projects'] = Project.objects.donated_projects(self.user_profile)
        statistics_dictionary = aggregate_stats(self.user_profile)
        statistics_dictionary['total_donated'] = total_donations(self.user_profile)
        total_people_affected = Project.objects.donated_completed_projects(self.user_profile)
        statistics_dictionary['people_served'] = total_people_affected
        humanize_integers(statistics_dictionary)
        admin_reinvestment = \
            Payment.objects.filter(user=self.user_profile).filter(admin_reinvestment__isnull=False).aggregate(
                Sum('amount'))[
                'amount__sum'] or 0
        user_reinvestment = UserReinvestment.objects.filter(user=self.user_profile).aggregate(Sum('amount'))[
                                'amount__sum'] or 0
        statistics_dictionary['reinvestment'] = admin_reinvestment + user_reinvestment
        reinvest_pool_amount = \
            RevolvUserProfile.objects.filter(reinvest_pool__isnull=False).aggregate(Sum('reinvest_pool'))[
            'reinvest_pool__sum']
        if reinvest_pool_amount > 0 & self.is_administrator:
            statistics_dictionary['reinvest_pool_amount'] = float("{0:.2f}".format(reinvest_pool_amount))
        else:
            statistics_dictionary['reinvest_pool_amount'] = 0
        context['statistics'] = statistics_dictionary

        return context


class CategoryPreferenceSetterView(UserDataMixin, View):
    http_methods = ['post']

    def http_method_not_allowed(request, *args, **kwargs):
        return redirect("dashboard")

    def post(self, request, *args, **kwargs):
        user = self.user_profile
        user.preferred_categories.clear()
        info_dict = request.POST.dict()
        for category_string in info_dict:
            category = Category.objects.get(id=category_string)
            user.preferred_categories.add(category)
        return HttpResponse()


class ProjectListView(UserDataMixin, TemplateView):
    """ Base View of all active projects
    """
    model = Project
    template_name = 'base/projects-list.html'

    def get_context_data(self, **kwargs):
        context = super(ProjectListView, self).get_context_data(**kwargs)
        active = Project.objects.get_active()
        context["active_projects"] = filter(lambda p: p.amount_left > 0.0, active)
        context["is_reinvestment"] = False
        return context

    def dispatch(self, request, *args, **kwargs):
        if request.GET:
            request.session["utm_params"] = request.GET

        return super(ProjectListView, self).dispatch(request, *args, **kwargs)

class SignInView(TemplateView):
    """Signup and login page. Has three submittable forms: login, signup,
    and sign in with facebook.

    Note that the signup with facebook functionality will automatically
    create a user object, but if the user has previously signed up with
    facebook, they will not be able to sign in without facebook to the
    same account.

    Also note that the "sign in with facebook" does not necessarily do
    the same thing as either sign up or login: when the user clicks this
    button, they will be automatically signed up and logged in if there
    is not an account associated with their facebook profile, or they will
    just be logged in if there is.

    Http verbs:
        GET: renders the signin page with empty forms.
    """
    template_name = "base/sign_in.html"
    login_form_class = SignInForm
    signup_form_class = SignupForm

    def dispatch(self, request, *args, **kwargs):
        amount = request.GET.get('donation_amount')
        tip = request.GET.get('donation_tip')
        title = request.GET.get('title')
        request.session['amount'] = amount
        request.session['tip'] = tip
        request.session['title'] = title
        if request.user.is_authenticated():
            return redirect("home")
        return super(SignInView, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, *args, **kwargs):
        context = super(SignInView, self).get_context_data(**kwargs)
        login_form = self.login_form_class()
        signup_form = self.signup_form_class()
        context["signup_form"] = signup_form
        context["login_form"] = login_form
        if self.request.GET.get("next"):
            context["login_redirect_url"] = self.request.GET.get("next")
        else:
            context["login_redirect_url"] = reverse('dashboard')
        context["referring_endpoint"] = ""
        context["reason"] = self.request.GET.get("reason")
        return context


class RedirectToSigninOrHomeMixin(object):
    """
    Mixin that detects if a page was requested with method="GET", and redirects
    to the signin page if so. Also, if posted to with an already authenticated
    user, will redirect to the homepage instead.

    This is useful both for the login and signup endpoints.
    """

    @method_decorator(sensitive_post_parameters(
        "password", "password1", "password2"
    ))
    def dispatch(self, request, *args, **kwargs):
        # don't allow rendering a form page with a GET request to this view:
        # instead, redirect to the signin page
        if self.request.method == "GET":
            return redirect("signin")
        # if the user is already logged in and tries to log in again, just
        # redirect them to the home page.
        if request.user.is_authenticated():
            return redirect("home")
        return super(RedirectToSigninOrHomeMixin, self).dispatch(
            request, *args, **kwargs
        )


class LoginView(RedirectToSigninOrHomeMixin, FormView):
    """
    Login endpoint: checks the data from the received request against
    django.contrib.auth.forms.AuthenticationForm and logs in the user if
    possible. If not, redirects back to the signin page.

    Http verbs:
        GET: redirect to signin page
        POST: check post parameters for user credentials, login the user
            and redirect to the specified next page (home by default), or
            render the sign in page with errors.
    """
    form_class = SignInForm
    template_name = 'base/sign_in.html'
    url_append = "#login"
    redirect_view = "signin"

    @csrf_exempt
    @method_decorator(sensitive_post_parameters('password'))
    def dispatch(self, request, *args, **kwargs):
        if request.COOKIES.get("last_project"):
            self.next_url = request.POST.get("next", "dashboard")
        else:
            self.next_url = request.POST.get("next", "home")
        return super(LoginView, self).dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        """Log the user in and redirect them to the supplied next page."""
        auth_login(self.request, form.get_user())
        if self.request.session.get('payment'):
            Payment.objects.filter(id=self.request.session['payment']).update(
                user_id=self.request.user.revolvuserprofile, entrant_id=self.request.user.revolvuserprofile)
            payment = Payment.objects.get(id=self.request.session['payment'])
            Tip.objects.filter(id=payment.tip_id).update(user_id=self.request.user.revolvuserprofile)
            Project.objects.get(id=payment.project_id).donors.add(self.request.user.revolvuserprofile)
            AnonymousUserDonation.objects.filter(payment_id=self.request.session['payment']).delete()
            del self.request.session['payment']

            # messages.success(self.request, 'Logged in as ' + self.request.POST.get('username'))
            # return redirect(reverse('project:view', kwargs={'title':title})+'?amount='+amount+'&tip='+tip)
        messages.success(self.request, 'Logged in as ' + self.request.POST.get('username'))
        return redirect(self.next_url)

    def get_context_data(self, *args, **kwargs):
        context = super(LoginView, self).get_context_data(*args, **kwargs)
        context["signup_form"] = SignupForm()
        context["login_form"] = self.get_form(self.form_class)
        context["referring_endpoint"] = "login"
        return context


class SignupView(RedirectToSigninOrHomeMixin, FormView):
    """
    Signup endpoint: processes the signup form and signs the user up (and logs
    them in). Note that the sign up with facebook functionality is entirely
    different: this is only for the regular django auth signup flow.

    Http verbs:
        GET: redirect to signin page
        POST: check post params against form, redirect to signin page if the
            form is not valid.
    """
    form_class = SignupForm
    template_name = "base/sign_in.html"
    url_append = "#signup"
    redirect_view = "signin"

    @csrf_exempt
    @method_decorator(sensitive_post_parameters('password'))
    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated():
            return redirect("dashboard")
        return super(SignupView, self).dispatch(request, *args, **kwargs)

    @csrf_exempt
    def form_valid(self, form):
        form.save()
        u = form.ensure_authenticated_user()
        name = u.revolvuserprofile.get_full_name()
        send_signup_info(name, u.email, u.revolvuserprofile.address)
        # log in the newly created user model. if there is a problem, error
        auth_login(self.request, u)
        SITE_URL = settings.SITE_URL
        login_link = SITE_URL + reverse('login')
        portfolio_link = SITE_URL + reverse('dashboard')
        context = {}
        context['user'] = self.request.user
        context['login_link'] = login_link
        context['portfolio_link'] = portfolio_link

        send_revolv_email(
            'signup',
            context, [self.request.user.email]
        )

        user = RevolvUserProfile.objects.get(user=self.request.user)
        if user.subscribed_to_newsletter:
            try:
                list = mailchimp.utils.get_connection().get_list_by_id(LIST_ID)
                list.con.list_subscribe(list.id, self.request.user.email,
                                        {'EMAIL': self.request.user.email,
                                         'FNAME': self.request.user.first_name,
                                         'LNAME': self.request.user.last_name,
                                         'INTERESTS': NEWSLETTERS},
                                        double_optin=False, update_existing=True)
            except Exception, e:
                logger.exception(e)

        if self.request.session.get('payment'):
            Payment.objects.filter(id=self.request.session['payment']).update(
                user_id=self.request.user.revolvuserprofile, entrant_id=self.request.user.revolvuserprofile)
            payment = Payment.objects.get(id=self.request.session['payment'])
            Tip.objects.filter(id=payment.tip_id).update(user_id=self.request.user.revolvuserprofile)
            Project.objects.get(id=payment.project_id).donors.add(self.request.user.revolvuserprofile)
            AnonymousUserDonation.objects.filter(payment_id=self.request.session['payment']).delete()
            del self.request.session['payment']
        messages.success(self.request, 'Signed up successfully!')
        return redirect("dashboard")

    def get_context_data(self, *args, **kwargs):
        context = super(SignupView, self).get_context_data(**kwargs)
        context["signup_form"] = self.get_form(self.form_class)
        context["login_form"] = SignInForm()
        context["referring_endpoint"] = "signup"
        return context


class LogoutView(UserDataMixin, View):
    """
    Basic logout view: Accessed whenever the user wants to logout, processes
    the logout, shows a toast, and redirects to home.
    """

    def get(self, request, *args, **kwargs):
        auth_logout(request)
        messages.success(self.request, 'Logged out successfully')
        return redirect('home')


def faq(request):
    return render(request, 'base/partials/faq.html')


def myths_and_facts(request):
    return render(request, 'base/partials/myth_and_facts.html')


def solarathome(request):
    return render_to_response('base/solar_at_home.html',
                              context_instance=RequestContext(request))


def completedproject(request):
    completed_projects = Project.objects.get_completed()
    # return render_to_response('base/partials/completed_projects.html',context_instance=RequestContext(request))
    return render(request, 'base/partials/completed_projects.html', {'completed_projects': completed_projects})


def bring_solar_tou_your_community(request):
    return render_to_response('base/bring_solar_at_community.html',
                              context_instance=RequestContext(request))


def intake_form_submit(request):
    try:
        email = request.GET.get('email')
        zipCode = request.GET.get('zipCode')
        signUp = request.GET.get('signUp')
        interest = request.GET.get('interest')
        heardSource = request.GET.get('heardSource')
        personalDesc = request.GET.get('personalDesc')
        leadDesc = request.GET.get('leadDesc')
        organisationName = request.GET.get('organisationName')
        organisationTaxId = request.GET.get('organisationTaxId')
        organisationAddress = request.GET.get('organisationAddress')
        billingAddress = request.GET.get('billingAddress')
        websiteName = request.GET.get('websiteName')
        phoneNumber = request.GET.get('phoneNumber')
        missionStatement = request.GET.get('missionStatement')
        orgStartYear = request.GET.get('orgStartYear')
        affiliation = request.GET.get('affiliation')
        solarProjNeed = request.GET.get('solarProjNeed')
        annualBudget = request.GET.get('annualBudget')
        checkOwnBuilding = request.GET.get('checkOwnBulding')
        orgBuildingYears = request.GET.get('orgBuildingYears')
        folkCounts = request.GET.get('folkCounts')
        buildingRoofYear = request.GET.get('buildingRoofYear')
        roofReplace = request.GET.get('roofReplace')
        electricityProvider = request.GET.get('electricityProvider')
        orgInterestBlock = request.GET.get('orgInterestBlock')

    except:
        logger.exception('Form values are not valid')
        return HttpResponseBadRequest('bad POST data')

    context = {}
    context['email'] = email
    context['zipCode'] = zipCode
    context['signUp'] = signUp
    context['interest'] = interest
    context['heardSource'] = heardSource
    context['personalDesc'] = personalDesc
    context['leadDesc'] = leadDesc
    context['organisationName'] = organisationName
    context['organisationTaxId'] = organisationTaxId
    context['organisationAddress'] = organisationAddress
    context['billingAddress'] = billingAddress
    context['websiteName'] = websiteName
    context['phoneNumber'] = phoneNumber
    context['missionStatement'] = missionStatement
    context['orgStartYear'] = orgStartYear
    context['affiliation'] = affiliation
    context['solarProjNeed'] = solarProjNeed
    context['annualBudget'] = annualBudget
    context['checkOwnBuilding'] = checkOwnBuilding
    context['orgBuildingYears'] = orgBuildingYears
    context['folkCounts'] = folkCounts
    context['buildingRoofYear'] = buildingRoofYear
    context['roofReplace'] = roofReplace
    context['electricityProvider'] = electricityProvider
    context['orgInterestBlock'] = orgInterestBlock
    send_revolv_email(
        'intake_form',
        context, ['info@re-volv.org']
    )

    return redirect('bring_solar_to_your_community')


def select_chapter(request, chapter):
    if 1 <= int(chapter) <= 12:
        return render_to_response("base/chapter{0}.html".format(chapter if int(chapter) != 1 else ''),
                                context_instance=RequestContext(request))
    else:
        return redirect('bring_solar_to_your_community')


def intake_form(request):
    return render_to_response('base/intake_form.html',
                              context_instance=RequestContext(request))


class ReinvestmentRedirect(UserDataMixin, TemplateView):
    '''
    Redirect user to the active project list whenever user has reinvestment amount.
    '''
    model = Project
    template_name = 'base/projects-list.html'

    def get_object(self):
        return self.model.objects.get(pk=self.request.user.pk)

    def get_context_data(self, **kwargs):
        context = super(ReinvestmentRedirect, self).get_context_data(**kwargs)
        active = Project.objects.get_active()
        context["active_projects"] = filter(lambda p: p.amount_left > 0.0, active)
        if self.user_profile:
            amount = self.user_profile.reinvest_pool + self.user_profile.solar_seed_fund_pool
        if self.user_profile and amount > 0.0:
            context["is_reinvestment"] = True
            context["reinvestment_amount"] = self.user_profile.reinvest_pool + self.user_profile.solar_seed_fund_pool

        return context


class DashboardRedirect(UserDataMixin, View):
    """
    Redirects user to appropriate dashboard. (e.g. Administrators automagically
    go to the /my-portfolio/admin endpoint)

    Redirects to home page if not authenticated.
    """

    def get(self, request, *args, **kwargs):
        # social = request.GET.get('social', '')
        amount = request.session.get('amount')
        project = request.session.get('project')
        social = request.session.get('social')
        url = request.session.get('url')
        cover_photo = request.session.get('cover_photo', '')
        request.session['amount'] = amount
        request.session['project'] = project
        request.session['cover_photo'] = cover_photo
        request.session['social'] = social
        request.session['url'] = url

        if self.is_authenticated:
            if self.request.session.get('payment'):
                Payment.objects.filter(id=self.request.session['payment']).update(
                    user_id=self.request.user.revolvuserprofile, entrant_id=self.request.user.revolvuserprofile)
                payment = Payment.objects.get(id=self.request.session['payment'])
                Tip.objects.filter(id=payment.tip_id).update(user_id=self.request.user.revolvuserprofile)
                Project.objects.get(id=payment.project_id).donors.add(self.request.user.revolvuserprofile)
                AnonymousUserDonation.objects.filter(payment_id=self.request.session['payment']).delete()
                del self.request.session['payment']

        if bool(request.GET) is False:
            if not self.is_authenticated:
                return redirect('home')
            if self.is_administrator:
                return redirect('administrator:dashboard')
            if self.is_ambassador:
                return redirect('ambassador:dashboard')
            return redirect(reverse('donor:dashboard'))

        elif social == 'donation':
            if not self.is_authenticated:
                return redirect('home')

            if self.is_administrator:
                return redirect('administrator:dashboard')

            if self.is_ambassador:
                return redirect(reverse('ambassador:dashboard'))
            if self.is_donor:
                return redirect(reverse('donor:dashboard'))

        else:
            if not self.is_authenticated:
                return redirect('home')

            if self.is_administrator:
                return redirect('administrator:dashboard')

            if self.is_ambassador:
                return redirect(reverse('ambassador:dashboard') + '?social=' + social)
            if self.is_donor:
                return redirect(reverse('donor:dashboard') + '?social=' + social)


# password reset/change views: thin wrappers around django's built in password
# reset views, but with our own templates
def password_reset_initial(request):
    return auth_views.password_reset(
        request,
        template_name="base/auth/forgot_password_initial.html",
        email_template_name="base/auth/forgot_password_email.html",
        from_email="support@re-volv.org"
    )


def password_change(request):
    return auth_views.password_change(
        request,
        template_name="base/auth/change_password.html",
        post_change_redirect="/my-portfolio/donor/?password_change_success",
    )


def password_reset_done(request):
    return auth_views.password_reset_done(request, template_name="base/auth/forgot_password_done.html")


def password_reset_confirm(request, *args, **kwargs):
    kwargs.update({"template_name": "base/auth/forgot_password_confirm.html"})
    return auth_views.password_reset_confirm(request, *args, **kwargs)


def password_reset_complete(request):
    return auth_views.password_reset_complete(request, template_name="base/auth/forgot_password_complete.html")


@login_required
@require_http_methods(['GET'])
def unsubscribe(request, action):
    """
    View handle unsubscribe email update
    """
    data = {}
    if action and action.lower() == 'updates':
        user_profile = request.user.revolvuserprofile
        user_profile.subscribed_to_repayment_notifications = False
        user_profile.save()
        data = {'msg': "You have successfully unsubscribed"}
    elif action and action.lower() == 'confirmation':
        return render_to_response('base/unsub_conf.html', context_instance=RequestContext(request, data))
    else:
        data = {'msg': 'Please specify which section you want to unsubscribe'}

    return render_to_response('base/minimal_message.html',
                              context_instance=RequestContext(request, data))


@login_required
def social_connection(request):
    """
    View handle my social connection page
    """
    backend_map = {'facebook': {'name': 'facebook', 'connected': False,
                                'dc_url': reverse('social:disconnect', kwargs={'backend': 'facebook'})},
                   'google': {'name': 'google-oauth2', 'connected': False,
                              'dc_url': reverse('social:disconnect', kwargs={'backend': 'google-oauth2'})}
                   }
    accounts = UserSocialAuth.objects.filter(user=request.user)

    for account in accounts:
        for k, v in backend_map.iteritems():
            if v['name'] == account.provider:
                backend_map[k]['connected'] = True

    return render_to_response('base/social_account.html',
                              context_instance=RequestContext(request, {'accounts': backend_map}))


def harborhouse(request):
    return redirect('/project/harborhouse/')


def leo_page(request):
    return render(request, 'base/partials/leo_page.html')


def revolv_accelerator(request):
    context = {}
    active_projects = Project.objects.get_active()
    context["active_projects"] = filter(lambda p: p.amount_left > 0.0, active_projects)
    return render_to_response('base/partials/revolv_accelerator.html',
                              context_instance=RequestContext(request, {'active_projects': active_projects}))


def leadership_circle(request):
    return render(request, 'base/partials/leadership-circle.html')


def riverrevitalizationfoundation(request):
    return redirect('/project/riverrevitalizationfoundation/')


def faithbaptistchurch(request):
    return redirect('/project/faithbaptistchurch/')


def campketcha(request):
    return redirect('/project/campketcha/')


def sendmail(request):
    revolv_user = get_object_or_404(RevolvUserProfile, user_id=request.user.id)
    projects = Project.objects.all()
    project_list = []
    for project in projects:
        for ambassador in project.ambassadors.all():
            if revolv_user == ambassador:
                project_list.append(project)
    context = {}
    context['project_list'] = project_list
    return render(request, 'base/ambassador_send_email.html',
                  context)


def send_donor_email(request):
    try:
        pk = request.POST['project']
        email_text = request.POST['email_text']
        email_subject = request.POST['email_subject_text']

    except KeyError:
        logger.exception('send_donor_email called without required POST data')
        return HttpResponseBadRequest('bad POST data')

    emails = []
    project = get_object_or_404(Project, pk=pk)
    payments = Payment.objects.filter(project=project).filter(admin_reinvestment_id__isnull=True).distinct('user_id')
    for payment in payments:
        email = payment.user.user.email
        if email:
            if email not in emails:
                emails.append(email)
    context = {}
    context['email_text'] = email_text
    context['email_subject'] = email_subject

    for email in emails:
        send_revolv_email(
            'donor_template',
            context, [email]
        )
    messages.success(request, 'Emails are sent successfully.')
    return redirect('sendmail')


def social_exception(request):
    has_social_exception = request.session.get('has_social_exception')
    if not has_social_exception:
        return redirect(reverse('home'))

    del request.session['has_social_exception']

    message = request.GET.get('message')
    return render_to_response('base/minimal_message.html',
                              context_instance=RequestContext(request, {'msg': message}))


class MatchingDonorsView(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/matching_donors.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(MatchingDonorsView, self).dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        # self.object = self.get_object()
        context = super(MatchingDonorsView, self).get_context_data(**kwargs)
        context['donors'] = ProjectMatchingDonors.objects.all()
        context['users'] = RevolvUserProfile.objects.all()
        context['projects'] = Project.objects.all()
        return self.render_to_response(context)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(MatchingDonorsView, self).get_context_data(**kwargs)
        context['donors'] = ProjectMatchingDonors.objects.all()
        context['users'] = RevolvUserProfile.objects.all()
        context['projects'] = Project.objects.all()
        return context


@login_required
def delete(request):
    pk = request.REQUEST['id']
    ProjectMatchingDonor = get_object_or_404(ProjectMatchingDonors, pk=pk)
    ProjectMatchingDonor.delete()
    return HttpResponse(json.dumps({'status': 'delete'}), content_type="application/json")


@login_required
def edit(request):
    pk = request.REQUEST['id']
    ProjectMatchingDonor = get_object_or_404(ProjectMatchingDonors, pk=pk)
    serialized_obj = serializers.serialize('json', [ProjectMatchingDonor, ])
    return HttpResponse(json.dumps({'ProjectMatchingDonor': serialized_obj}), content_type="application/json")


@login_required
@require_http_methods(['POST'])
def add_maching_donors(request):
    id = request.POST.get('id')
    user = request.POST['user']
    project = request.POST['project']
    amount = request.POST['amount']
    revolv_user = get_object_or_404(RevolvUserProfile, pk=user)

    matching_project = get_object_or_404(Project, pk=project)

    if not id:
        ProjectMatchingDonors.objects.create(
            matching_donor=revolv_user,
            project=matching_project,
            amount=amount
        )
    else:
        projectMatchingDonor = ProjectMatchingDonors.objects.get(id=id)
        projectMatchingDonor.matching_donor = revolv_user
        projectMatchingDonor.project = matching_project
        projectMatchingDonor.amount = amount
        projectMatchingDonor.save()

    return HttpResponse(json.dumps({'status': 'created'}), content_type="application/json")


def ambassador_data_table(request):
    draw = request.GET.get('draw')
    datepicker1 = request.GET.get('datepicker1')
    datepicker2 = request.GET.get('datepicker2')
    length = request.GET.get('length')
    order = request.GET.get('order[0][dir]')
    start = request.GET.get('start')
    search = request.GET.get('search[value]')
    currentSortByCol = request.GET.get('order[0][column]')

    fields = ['user__user__first_name', 'user__user__last_name', 'user__user__username', 'user__user__email',
              'created_at', 'project__title', 'amount', 'user_reinvestment__amount', 'admin_reinvestment__amount',
              'tip__amount']

    order_by = {'desc': '-', 'asc': ''}

    column_order = order_by[order] + fields[int(currentSortByCol)]

    revolv_user = get_object_or_404(RevolvUserProfile, user_id=request.user.id)

    projects = Project.objects.all()
    project_list = []
    for project in projects:
        for ambassador in project.ambassadors.all():
            if revolv_user == ambassador:
                project_list.append(project)

    payment_list = Payment.objects.filter(project__in=project_list, admin_reinvestment__isnull=True).order_by((column_order))
    total_records_count = payment_list.count()
    if int(length) == -1:
        length = payment_list.count()
    if search.strip():
        payment_list = payment_list.filter(Q(user__user__username__icontains=search) |
                                           Q(user__user__first_name__icontains=search) |
                                           Q(project__title__icontains=search) |
                                           Q(amount__icontains=search) |
                                           Q(user__user__last_name__icontains=search) |
                                           Q(user__user__email__icontains=search))

    if datepicker1 and datepicker2:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(datepicker1, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(datepicker2, '%Y-%m-%d').date()

        payment_list = payment_list.filter(
            created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC),
                               datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0, pytz.UTC)])

    payments = []
    for payment in payment_list[int(start):][:int(length)]:

        payment_details_all = {}
        payment_details_all['firstname'] = payment.user.user.first_name
        payment_details_all['lastname'] = payment.user.user.last_name
        payment_details_all['username'] = payment.user.user.username
        payment_details_all['email'] = payment.user.user.email
        payment_details_all['date'] = (payment.created_at).strftime("%Y/%m/%d %H:%M:%S")
        payment_details_all['project'] = payment.project.title
        payment_details_all['amount'] = payment.amount
        if payment.tip:
            payment_details_all['tip'] = payment.tip.amount
        else:
            payment_details_all['tip'] = 0
        if payment.user_reinvestment:
            payment_details_all['amount'] = 0
            payment_details_all['tip'] = 0
            payment_details_all['user_reinvestment'] = round(payment.user_reinvestment.amount, 2)
        else:
            payment_details_all['user_reinvestment'] = 0
        total = payment_details_all['amount'] + payment_details_all['user_reinvestment'] + payment_details_all['tip']
        payment_details_all['total'] = total
        payments.append(payment_details_all)

    json_response = {"draw": draw, "recordsTotal": total_records_count,
                     "recordsFiltered": payment_list.count(), "all-data": payments}

    return HttpResponse(json.dumps(json_response), content_type='application/json')


def ambassador_data_table_auto_reinvestors(request):
    draw = request.GET.get('draw')
    datepicker1 = request.GET.get('datepicker1')
    datepicker2 = request.GET.get('datepicker2')
    length = request.GET.get('length')
    order = request.GET.get('order[0][dir]')
    start = request.GET.get('start')
    search = request.GET.get('search[value]')
    currentSortByCol = request.GET.get('order[0][column]')

    fields = ['user__user__first_name', 'user__user__last_name', 'user__user__username', 'user__user__email',
              'created_at', 'project__title', 'amount', 'user_reinvestment__amount', 'admin_reinvestment__amount',
              'tip__amount']

    order_by = {'desc': '-', 'asc': ''}

    column_order = order_by[order] + fields[int(currentSortByCol)]

    revolv_user = get_object_or_404(RevolvUserProfile, user_id=request.user.id)

    projects = Project.objects.all()
    project_list = []
    for project in projects:
        for ambassador in project.ambassadors.all():
            if revolv_user == ambassador:
                project_list.append(project)

    auto_reinvestors_list = Payment.objects.filter(project__in=project_list, admin_reinvestment__isnull=False, amount__gt=0).order_by((column_order))
    total_records_count = auto_reinvestors_list.count()
    if int(length) == -1:
        length = len(auto_reinvestors_list)
    if search.strip():
        auto_reinvestors_list = auto_reinvestors_list.filter(Q(user__user__username__icontains=search) |
                                           Q(user__user__first_name__icontains=search) |
                                           Q(project__title__icontains=search) |
                                           Q(amount__icontains=search) |
                                           Q(user__user__last_name__icontains=search) |
                                           Q(user__user__email__icontains=search))

    if datepicker1 and datepicker2:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(datepicker1, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(datepicker2, '%Y-%m-%d').date()

        auto_reinvestors_list = auto_reinvestors_list.filter(
            created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC),
                               datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0, pytz.UTC)])

    auto_reinvestors = []

    if auto_reinvestors_list:
        for payment in auto_reinvestors_list[int(start):][:int(length)]:
            if payment.amount > 0:
                payment_details = {}
                payment_details['firstname'] = payment.user.user.first_name
                payment_details['lastname'] = payment.user.user.last_name
                payment_details['username'] = payment.user.user.username
                payment_details['email'] = payment.user.user.email
                payment_details['date'] = (payment.created_at).strftime("%Y/%m/%d %H:%M:%S")
                payment_details['project'] = payment.project.title
                payment_details['admin_reinvestment'] = round(payment.amount, 2)
                auto_reinvestors.append(payment_details)
    json_response = {"draw": draw, "recordsTotal": total_records_count,
                     "recordsFiltered": len(auto_reinvestors_list), "auto-reinvestors-data": auto_reinvestors}

    return HttpResponse(json.dumps(json_response), content_type='application/json')


def payment_data_table(request):
    draw = request.GET.get('draw')
    date_picker_1 = request.GET.get('datepicker1')
    date_picker_2 = request.GET.get('datepicker2')
    length = request.GET.get('length')
    order = request.GET.get('order[0][dir]')
    start = request.GET.get('start')
    search = request.GET.get('search[value]').strip()
    current_sort_by_col = request.GET.get('order[0][column]')
    report_length = request.GET.get('reportLength')

    fields = ['user__user__first_name', 'user__user__last_name', 'user__user__username', 'user__user__email',
              'created_at', 'project__title', 'amount', 'user_reinvestment__amount', 'admin_reinvestment__amount',
              'tip__amount']

    order_by = {'desc': '-', 'asc': ''}

    column_order = order_by[order] + fields[int(current_sort_by_col)]

    auth_user = get_object_or_404(RevolvUserProfile, user_id=request.user.id)

    projects = Project.objects.all()
    project_list = []
    if request.user.revolvuserprofile.is_administrator():
        for project in projects:
            project_list.append(project)
    elif request.user.revolvuserprofile.is_ambassador():
        for project in projects:
            for ambassador in project.ambassadors.all():
                if auth_user == ambassador:
                    project_list.append(project)
    else:
        return HttpResponseRedirect(reverse("dashboard"))

    payment_list = Payment.objects.filter(project__in=project_list).order_by(column_order)
    total_records_count = payment_list.count()
    if int(report_length) == -4:
        payment_list = Payment.objects.filter(project__in=project_list,
                                              admin_reinvestment__isnull=True,
                                              user_reinvestment__isnull=True).order_by(column_order)
        total_records_count = payment_list.count()
    if int(report_length) == -3:
        payment_list = Payment.objects.filter(project__in=project_list,
                                              admin_reinvestment__isnull=True).order_by(
            column_order)
        total_records_count = payment_list.count()
    report_length = int(report_length)
    if int(report_length) < 0:
        if -2 < int(report_length) or int(report_length) < -4 or not isinstance(report_length, int):
            return HttpResponseRedirect(reverse("donationreport"))
        length = 10
    if search:
        anonymous_donations = AnonymousUserDonation.objects.\
            filter(email__icontains=search).\
            values_list('payment_id', flat=True)
        payment_list = payment_list.filter(Q(user__user__username__icontains=search) |
                                           Q(user__user__first_name__icontains=search) |
                                           Q(project__title__icontains=search) |
                                           Q(id__in=anonymous_donations) |
                                           Q(amount__icontains=search) |
                                           Q(user__user__last_name__icontains=search) |
                                           Q(user__user__email__icontains=search))

    if date_picker_1 and date_picker_2:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(date_picker_1, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(date_picker_2, '%Y-%m-%d').date()

        payment_list = payment_list.filter(
            created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC),
                               datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0, pytz.UTC)])

    payments = []

    for payment in payment_list[int(start):][:int(length)]:

        anonymous_donor_filter = AnonymousUserDonation.objects.filter(payment_id=payment.id)
        payment_details = dict()
        payment_details['first_name'] = payment.user.user.first_name
        payment_details['last_name'] = payment.user.user.last_name
        payment_details['username'] = payment.user.user.username
        if anonymous_donor_filter:
            payment_details['email'] = AnonymousUserDonation.objects.get(payment_id=payment.id).email
        else:
            payment_details['email'] = payment.user.user.email
        payment_details['date'] = payment.created_at.strftime("%Y/%m/%d %H:%M:%S")
        payment_details['project'] = payment.project.title
        if payment.user_reinvestment or payment.admin_reinvestment or payment.project.title == "Operations":
            payment_details['amount'] = 0
        else:
            payment_details['amount'] = payment.amount
        if payment.user_reinvestment:
            payment_details['user_reinvestment'] = round(payment.user_reinvestment.amount, 2)
        else:
            payment_details['user_reinvestment'] = 0
        if payment.admin_reinvestment:
            payment_details['admin_reinvestment'] = round(payment.amount, 2)
        else:
            payment_details['admin_reinvestment'] = 0
        if payment.tip and payment.project.title == "Operations":
            payment_details['tip'] = round(payment.tip.amount + payment.amount, 2)
        if payment.project.title == "Operations":
            payment_details['tip'] = round(payment.amount, 2)
        elif payment.tip:
            payment_details['tip'] = round(payment.tip.amount, 2)
        else:
            payment_details['tip'] = 0
        if payment.tip and payment.amount:
            payment_details['total'] = round(payment.tip.amount + payment.amount, 2)
        if payment.tip and not payment.amount:
            payment_details['total'] = round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            payment_details['total'] = round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            payment_details['total'] = 0
        payments.append(payment_details)

    json_response = {"draw": draw, "recordsTotal": total_records_count,
                     "recordsFiltered": payment_list.count(), "data": payments}

    return HttpResponse(json.dumps(json_response), content_type='application/json')


def repayment_table(request):
    draw = request.GET.get('draw')
    datepicker1 = request.GET.get('datepicker1')
    datepicker2 = request.GET.get('datepicker2')
    length = request.GET.get('length')
    order = request.GET.get('order[0][dir]')
    start = request.GET.get('start')
    search = request.GET.get('search[value]')
    currentSortByCol = request.GET.get('order[0][column]')

    fields = ['user__user__first_name', 'user__user__last_name', 'user__user__username', 'user__user__email',
              'created_at', 'amount', 'admin_repayment__amount', 'project__title']

    order_by = {'desc': '-', 'asc': ''}

    column_order = order_by[order] + fields[int(currentSortByCol)]

    repayment_list = RepaymentFragment.objects.filter(amount__gt=0.00).order_by((column_order))

    if search.strip():
        repayment_list = repayment_list.filter(Q(user__user__username__icontains=search) |
                                               Q(user__user__first_name__icontains=search) |
                                               Q(project__title__icontains=search) |
                                               Q(amount__icontains=search) |
                                               Q(user__user__last_name__icontains=search) |
                                               Q(user__user__email__icontains=search))

    if datepicker1 and datepicker2:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(datepicker1, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(datepicker2, '%Y-%m-%d').date()

        repayment_list = repayment_list.filter(
            created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC),
                               datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0, pytz.UTC)])

    payments = []

    for payment in repayment_list[int(start):][:int(length)]:
        payment_details = {}
        payment_details['firstname'] = payment.user.user.first_name
        payment_details['lastname'] = payment.user.user.last_name
        payment_details['username'] = payment.user.user.username
        payment_details['email'] = payment.user.user.email
        payment_details['date'] = (payment.created_at).strftime("%Y/%m/%d %H:%M:%S")
        payment_details['project'] = payment.project.title
        payment_details['donated_amount'] = round(
            Payment.objects.filter(user=payment.user).filter(project=payment.project).aggregate(Sum('amount'))[
                'amount__sum'] or 0, 2)
        payment_details['repayment_amount'] = round(payment.amount, 2)

        payments.append(payment_details)

    json_response = {"draw": draw, "recordsTotal": RepaymentFragment.objects.all().count(),
                     "recordsFiltered": repayment_list.count(), "data": payments}

    return HttpResponse(json.dumps(json_response), content_type='application/json')


import csv
from StringIO import StringIO

from django.http import StreamingHttpResponse


def export_csv(request):
    from django.utils.encoding import smart_str
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    search = request.GET.get('search_value') or ''
    report_length = request.GET.get('report_length')

    auth_user = get_object_or_404(RevolvUserProfile, user_id=request.user.id)

    projects = Project.objects.all()
    project_list = []
    if request.user.revolvuserprofile.is_administrator():
        for project in projects:
            project_list.append(project)
    elif request.user.revolvuserprofile.is_ambassador():
        for project in projects:
            for ambassador in project.ambassadors.all():
                if auth_user == ambassador:
                    project_list.append(project)
    else:
        return HttpResponseRedirect(reverse("dashboard"))

    payment_list = Payment.objects.filter(project__in=project_list).\
        select_related("user", "project", "admin_reinvestment", "user_reinvestment", "tip", "user__user")
    if int(report_length) == -4:
        payment_list = Payment.objects.filter(project__in=project_list,
                                              admin_reinvestment__isnull=True,
                                              user_reinvestment__isnull=True).\
        select_related("user", "project", "admin_reinvestment", "user_reinvestment", "tip", "user__user")
    if int(report_length) == -3:
        payment_list = Payment.objects.filter(project__in=project_list,
                                              admin_reinvestment__isnull=True).\
        select_related("user", "project", "admin_reinvestment", "user_reinvestment", "tip", "user__user")
    anonymous_donations = AnonymousUserDonation.objects.filter(email__icontains=search).\
        values_list('payment_id', flat=True)
    if search:
        payment_list = payment_list.filter(Q(user__user__username__icontains=search) |
                                           Q(user__user__first_name__icontains=search) |
                                           Q(project__title__icontains=search) |
                                           Q(id__in=anonymous_donations) |
                                           Q(amount__icontains=search) |
                                           Q(user__user__last_name__icontains=search) |
                                           Q(user__user__email__icontains=search))

    if int(report_length) < 0:
        if -2 < int(report_length) or int(report_length) < -4:
            return HttpResponseRedirect(reverse("donationreport"))
    if from_date and to_date:
        import datetime
        import pytz

        date1 = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()

        payment_list = payment_list.filter(
            created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC),
                               datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0, pytz.UTC)])

    # Define a generator to stream data directly to the client
    def stream():
        buffer_ = StringIO()
        writer = csv.writer(buffer_)
        writer.writerow([
            smart_str(u"FIRST NAME"),
            smart_str(u"LAST NAME"),
            smart_str(u"USERNAME"),
            smart_str(u"EMAIL"),
            smart_str(u"DATE"),
            smart_str(u"NAME OF PROJECT"),
            smart_str(u"DONATION TO SOLAR SEED FUND"),
            smart_str(u"REINVESTMENT IN SOLAR SEED FUND"),
            smart_str(u"ADMIN REINVESTMENT IN SOLAR SEED FUND"),
            smart_str(u"DONATION TO OPERATION"),
            smart_str(u"TOTAL DONATIONS"),

        ])
        for payment in payment_list:
            user_info = payment.user.user
            if payment.admin_reinvestment:
                admin_reinvestment = round(payment.amount, 2)
            else:
                admin_reinvestment = 0

            if payment.user_reinvestment:
                user_reinvestment = round(payment.user_reinvestment.amount, 2)
            else:
                user_reinvestment = 0

            if payment.admin_reinvestment or payment.user_reinvestment:
                donation_amount = 0
            else:
                donation_amount = payment.amount

            if payment.tip:
                tip = round(payment.tip.amount, 2)
            else:
                tip = 0

            if payment.tip and payment.amount:
                total = round(payment.tip.amount + payment.amount, 2)
            if payment.tip and not payment.amount:
                total = round(payment.tip.amount, 2)
            if payment.amount and not payment.tip:
                total = round(payment.amount, 2)
            if not payment.amount and not payment.tip:
                total = 0
            if AnonymousUserDonation.objects.filter(payment_id=payment.id):
                email = AnonymousUserDonation.objects.get(payment_id=payment.id).email
            else:
                email = payment.user.user.email

            writer.writerow([
                smart_str(user_info.first_name),
                smart_str(user_info.last_name),
                smart_str(user_info.username),
                smart_str(email),
                smart_str(payment.created_at),
                smart_str(payment.project.title),
                smart_str(donation_amount),
                smart_str(user_reinvestment),
                smart_str(admin_reinvestment),
                smart_str(tip),
                smart_str(total),
            ])
            buffer_.seek(0)
            data = buffer_.read()
            buffer_.seek(0)
            buffer_.truncate()
            yield data
    # Create the streaming response  object with the appropriate CSV header.
    response = StreamingHttpResponse(stream(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="RE-volv_report.csv"'
    return response


def export_xlsx(request):
    """
       Export financial report Excel from the admin side.
       :param request:
       :return: Generate Excel report of selected records.
       """
    import openpyxl
    try:
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    search = request.GET.get('search_value') or ''
    search_query = Q()
    if search:
        search_query = Q(user__user__username__icontains=search) | \
                       Q(user__user__first_name__icontains=search) | \
                       Q(project__title__icontains=search) | \
                       Q(amount__icontains=search) | \
                       Q(user__user__last_name__icontains=search) | \
                       Q(user__user__email__icontains=search)
    if from_date and to_date:
        import datetime
        import pytz
        date1 = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        payments = Payment.objects.filter(
            created_at__range=[datetime.datetime(date1.year, date1.month, date1.day, 8, 15, 12, 0, pytz.UTC),
                               datetime.datetime(date2.year, date2.month, date2.day, 8, 15, 12, 0,
                                                 pytz.UTC)]).order_by('-created_at').filter(search_query)\
            .select_related("user", "project", "admin_reinvestment", "user_reinvestment", "tip", "user__user").iterator()
    else:
        payments = Payment.objects.filter(search_query).order_by('-created_at') \
            .select_related("user", "project", "admin_reinvestment", "user_reinvestment", "tip", "user__user")\
            .iterator()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=RE-volv.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "RE-volv"

    row_num = 0

    columns = [
        (u"FIRST NAME", 30),
        (u"LAST NAME", 30),
        (u"USERNAME", 30),
        (u"EMAIL", 30),
        (u"DATE", 30),
        (u"NAME OF PROJECT", 30),
        (u"DONATION TO SOLAR SEED FUND", 30),
        (u"REINVESTMENT IN SOLAR SEED FUND", 20),
        (u"ADMIN REINVESTMENT IN SOLAR SEED FUND", 20),
        (u"DONATION TO OPERATION", 20),
        (u"TOTAL DONATIONS", 20),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for payment in payments:
        if payment.admin_reinvestment:
            admin_reinvestment = round(payment.amount, 2)
        else:
            admin_reinvestment = 0

        if payment.user_reinvestment:
            user_reinvestment = round(payment.user_reinvestment.amount, 2)
        else:
            user_reinvestment = 0

        if payment.admin_reinvestment or payment.user_reinvestment:
            donation_amount = 0
        else:
            donation_amount = payment.amount

        if payment.tip:
            tip = round(payment.tip.amount, 2)
        else:
            tip = 0

        if payment.tip and payment.amount:
            total = round(payment.tip.amount + payment.amount, 2)
        if payment.tip and not payment.amount:
            total = round(payment.tip.amount, 2)
        if payment.amount and not payment.tip:
            total = round(payment.amount, 2)
        if not payment.amount and not payment.tip:
            total = 0
        if AnonymousUserDonation.objects.filter(payment_id=payment.id):
            email = AnonymousUserDonation.objects.get(payment_id=payment.id).email
        else:
            email = payment.user.user.email

        row_num += 1
        row = [
            payment.user.user.first_name,
            payment.user.user.last_name,
            payment.user.user.username,
            email,
            payment.created_at,
            payment.project.title,
            donation_amount,
            user_reinvestment,
            admin_reinvestment,
            tip,
            total,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]

    wb.save(response)
    payments.close()
    return response


def export_repayment_csv(request):
    """
    Export financial report CSV from the admin side.
    :param request:
    :return: Generate CSV report of selected records.
    """
    import csv
    from django.utils.encoding import smart_str
    # response = HttpResponse(content_type='text/csv')
    # response['Content-Disposition'] = 'attachment; filename=Repayment_report.csv'
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    search = request.GET.get('search_value') or ''
    search_query = Q()
    if search:
        search_query = Q(user__user__username__icontains=search) | \
                       Q(user__user__first_name__icontains=search) | \
                       Q(project__title__icontains=search) | \
                       Q(amount__icontains=search) | \
                       Q(user__user__last_name__icontains=search) | \
                       Q(user__user__email__icontains=search)
    if from_date and to_date:
        import datetime
        import pytz
        date1 = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        repayments = RepaymentFragment.objects.filter(amount__gt=0.00,
                                                      created_at__range=[
                                                          datetime.datetime(date1.year, date1.month, date1.day, 8, 15,
                                                                            12, 0, pytz.UTC),
                                                          datetime.datetime(date2.year, date2.month, date2.day, 8, 15,
                                                                            12, 0, pytz.UTC)]).order_by('-created_at') \
            .select_related("user", "project", "user__user").filter(search_query).iterator()
    else:
        repayments = RepaymentFragment.objects.filter(amount__gt=0.00).filter(search_query).order_by('-created_at') \
            .select_related("user", "project", "user__user").iterator()
    # writer = csv.writer(response, csv.excel)
    # response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)

    def stream():
        buffer_ = StringIO()
        writer = csv.writer(buffer_)
        writer.writerow([
            smart_str(u"FIRST NAME"),
            smart_str(u"LAST NAME"),
            smart_str(u"USERNAME"),
            smart_str(u"EMAIL"),
            smart_str(u"DATE"),
            smart_str(u"NAME OF PROJECT"),
            smart_str(u"DONATION AMOUNT"),
            smart_str(u"REPAYMENT AMOUNT"),

        ])

        for payment in repayments:
            writer.writerow([
                smart_str(payment.user.user.first_name),
                smart_str(payment.user.user.last_name),
                smart_str(payment.user.user.username),
                smart_str(payment.user.user.email),
                smart_str(payment.created_at),
                smart_str(payment.project.title),
                smart_str(round(
                    Payment.objects.filter(user=payment.user).filter(project=payment.project).aggregate(Sum('amount'))[
                        'amount__sum'] or 0, 2)),
                smart_str(round(payment.amount, 2)),
            ])
            buffer_.seek(0)
            data = buffer_.read()
            buffer_.seek(0)
            buffer_.truncate()
            yield data

            # Create the streaming response  object with the appropriate CSV header.
    response = StreamingHttpResponse(stream(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Repayment_report.csv"'
    return response


def export_repayment_xlsx(request):
    """
       Export financial report Excel from the admin side.
       :param request:
       :return: Generate Excel report of selected records.
       """
    import openpyxl
    try:
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    search = request.GET.get('search_value') or ''
    search_query = Q()
    if search:
        search_query = Q(user__user__username__icontains=search) | \
                       Q(user__user__first_name__icontains=search) | \
                       Q(project__title__icontains=search) | \
                       Q(amount__icontains=search) | \
                       Q(user__user__last_name__icontains=search) | \
                       Q(user__user__email__icontains=search)
    if from_date and to_date:
        import datetime
        import pytz
        date1 = datetime.datetime.strptime(from_date, '%Y-%m-%d').date()
        date2 = datetime.datetime.strptime(to_date, '%Y-%m-%d').date()
        repayments = RepaymentFragment.objects.filter(amount__gt=0.00,
                                                      created_at__range=[
                                                          datetime.datetime(date1.year, date1.month, date1.day, 8, 15,
                                                                            12, 0, pytz.UTC),
                                                          datetime.datetime(date2.year, date2.month, date2.day, 8, 15,
                                                                            12, 0, pytz.UTC)]).order_by('-created_at') \
            .select_related("user", "project", "user__user").filter(search_query).iterator()
    else:
        repayments = RepaymentFragment.objects.filter(amount__gt=0.00).order_by('-created_at') \
            .select_related("user", "project", "user__user").filter(search_query).iterator()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Repayment_report.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "RE-volv"

    row_num = 0

    columns = [
        (u"FIRST NAME", 30),
        (u"LAST NAME", 30),
        (u"USERNAME", 30),
        (u"EMAIL", 30),
        (u"DATE", 30),
        (u"NAME OF PROJECT", 30),
        (u"DONATION AMOUNT", 30),
        (u"REPAYMENT AMOUNT", 30),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for payment in repayments:
        row_num += 1
        row = [
            payment.user.user.first_name,
            payment.user.user.last_name,
            payment.user.user.username,
            payment.user.user.email,
            payment.created_at,
            payment.project.title,
            round(Payment.objects.filter(user=payment.user).filter(project=payment.project).aggregate(Sum('amount'))[
                      'amount__sum'] or 0, 2),
            round(payment.amount, 2)

        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    wb.save(response)
    return response


def add_email_to_mailing_list(request):
    if request.POST['email']:
        email_address = request.POST['email']
        list = mailchimp.utils.get_connection().get_list_by_id(LIST_ID)
        try:
            list.con.list_subscribe(list.id, email_address,
                                    {'EMAIL': email_address,
                                     'INTERESTS': NEWSLETTERS},
                                    double_optin=False, update_existing=True)
        except Exception:
            return HttpResponse(json.dumps({'status': 'subscription_fail'}), content_type="application/json")

        return HttpResponse(json.dumps({'status': 'subscription_success'}), content_type="application/json")
    else:
        return HttpResponse(json.dumps({'status': 'subscription_fail'}), content_type="application/json")


class editprofile(View):
    def post(self, request):
        subscribed_to_newsletter = request.POST.get('newsletter')
        repayment_notification = request.POST.get('repayment_notification')
        announcement = request.POST.get('announcement')
        profileup = RevolvUserProfileForm(data=request.POST or None, instance=request.user.revolvuserprofile)
        if (profileup.is_valid()) and (request.user.email == request.POST.get('email')):
            user = profileup.save(commit=False)

            interests = None
            if subscribed_to_newsletter:
                user.subscribed_to_newsletter = True
                interests = NEWSLETTERS

            if announcement:
                user.subscribed_to_updates = True
                if interests:
                    interests = interests + ", " + ANNOUNCEMENTS
                else:
                    interests = ANNOUNCEMENTS

            list = mailchimp.utils.get_connection().get_list_by_id(LIST_ID)

            if interests:
                list.con.list_subscribe(list.id, request.user.email, {'EMAIL': request.user.email,
                                                                      'FNAME': request.user.first_name,
                                                                      'LNAME': request.user.last_name,
                                                                      'INTERESTS': interests},
                                        double_optin=False, update_existing=True)
            else:
                try:
                    list.con.list_unsubscribe(list.id, request.user.email, delete_member=True)
                except:
                    logger.debug("Error while unsubscribe")
            if repayment_notification:
                user.subscribed_to_repayment_notifications = True

            user.user = request.user
            user.save()
            userup = UpdateUser(request.POST or None, instance=request.user)
            if userup.is_valid():
                user = userup.save(commit=False)
                user.save()
                messages.success(request, 'Account details successfully updated')
                return HttpResponseRedirect('/account_settings/')

            else:
                existing_user = False
                userprofile = RevolvUserProfile.objects.get(user=request.user)
                project = Project.objects.get(title='Operations')
                donated_solar_seed = \
                    Payment.objects.filter(user=userprofile).exclude(project=project).aggregate(Sum('amount'))[
                        'amount__sum'] or 0
                repayment_solar_seed = RepaymentFragment.objects.filter(user=userprofile).aggregate(Sum('amount'))[
                                           'amount__sum'] or 0
                operation_donation = \
                    Payment.objects.filter(user=userprofile, project=project).aggregate(Sum('amount'))[
                        'amount__sum'] or 0
                monthly_operation_donation = StripeDetails.objects.filter(user=userprofile).filter(amount__gt=0.0)
                monthly_solar_donation = StripeDetails.objects.filter(user=userprofile).filter(donation_amount__gt=0.0)
                monthly_donation_amount = 0.0
                solar_donation = 0.0
                if monthly_operation_donation or monthly_solar_donation:
                    existing_user = True
                    operation_amount = monthly_operation_donation.aggregate(Sum('amount'))['amount__sum'] or 0.0
                    solar_amount = monthly_solar_donation.aggregate(Sum('donation_amount'))[
                                       'donation_amount__sum'] or 0.0
                    monthly_donation_amount = operation_amount
                    solar_donation = solar_amount
                context = {
                    'subscribed_to_newsletter': userprofile.subscribed_to_newsletter,
                    'subscribed_to_repayment_notifications': userprofile.subscribed_to_repayment_notifications,
                    'subscribed_to_updates': userprofile.subscribed_to_updates,
                    'donated_solar_seed': donated_solar_seed,
                    'repayment_solar_seed': repayment_solar_seed,
                    'operation_donation': operation_donation,
                    'monthly_donation_amount': monthly_donation_amount,
                    'monthly_solar_donation': solar_donation,
                    'existing_user': existing_user,
                    "form": userup
                }
                return render(request, 'base/partials/account_settings.html', context)
        else:
            messages.error(request, 'Account details update failed.')
            return HttpResponseRedirect('/account_settings/')


@login_required
def account_settings(request):
    existing_user = False
    user = request.user
    userprofile = RevolvUserProfile.objects.get(user=request.user)
    project = Project.objects.get(title='Operations')
    donated_solar_seed = Payment.objects.filter(user=userprofile).exclude(project=project).aggregate(Sum('amount'))[
                             'amount__sum'] or 0
    repayment_solar_seed = RepaymentFragment.objects.filter(user=userprofile).aggregate(Sum('amount'))[
                               'amount__sum'] or 0
    operation_donation = Payment.objects.filter(user=userprofile, project=project).aggregate(Sum('amount'))[
                             'amount__sum'] or 0
    tip = Tip.objects.filter(user=userprofile).aggregate(Sum('amount'))['amount__sum'] or 0
    userform = UpdateUser(
        initial={'first_name': user.first_name, 'last_name': user.last_name, 'username': user.username,
                 'email': user.email})
    revolv_profile = RevolvUserProfile.objects.get(user=request.user)

    monthly_operation_donation = StripeDetails.objects.filter(user=revolv_profile).filter(amount__gt=0.0)
    monthly_solar_donation = StripeDetails.objects.filter(user=revolv_profile).filter(donation_amount__gt=0.0)
    monthly_donation_amount = 0.0
    solar_donation = 0.0
    if monthly_operation_donation or monthly_solar_donation:
        existing_user = True
        operation_amount = monthly_operation_donation.aggregate(Sum('amount'))['amount__sum'] or 0.0
        solar_amount = monthly_solar_donation.aggregate(Sum('donation_amount'))['donation_amount__sum'] or 0.0
        monthly_donation_amount = operation_amount
        solar_donation = solar_amount
    context = {
        'subscribed_to_newsletter': userprofile.subscribed_to_newsletter,
        'subscribed_to_repayment_notifications': userprofile.subscribed_to_repayment_notifications,
        'subscribed_to_updates': userprofile.subscribed_to_updates,
        'donated_solar_seed': donated_solar_seed,
        'repayment_solar_seed': repayment_solar_seed,
        'operation_donation': operation_donation + tip,
        'monthly_donation_amount': monthly_donation_amount,
        'monthly_solar_donation': solar_donation,
        'existing_user': existing_user,
        "form": userform
    }
    return render(request, 'base/partials/account_settings.html', context)


def create_subscription(request, donation_type, revolv_profile, amt_in_cents, customer):
    if donation_type == 'SOLAR_SEED_FUND':
        plan = stripe.Plan.create(
            amount=int(amt_in_cents),
            interval="month",
            name="Solar Donation " + str(amt_in_cents / 100),
            currency="usd",
            id="solar_donation" + "_" + customer + "_" + str(amt_in_cents / 100))

        subscription = stripe.Subscription.create(
            customer=customer,
            plan=plan,
        )
        StripeDetails.objects.create(
            user=revolv_profile,
            stripe_customer_id=subscription.customer,
            subscription_id=subscription.id,
            plan=subscription.plan.id,
            stripe_email=request.user.email,
            donation_amount=amt_in_cents / 100
        )

    else:
        plan = stripe.Plan.create(
            amount=int(amt_in_cents),
            interval="month",
            name="Operation Donation " + str(amt_in_cents / 100),
            currency="usd",
            id="operation_donation" + "_" + customer + "_" + str(amt_in_cents / 100))

        subscription = stripe.Subscription.create(
            customer=customer,
            plan=plan,
        )
        StripeDetails.objects.create(
            user=revolv_profile,
            stripe_customer_id=subscription.customer,
            subscription_id=subscription.id,
            plan=subscription.plan.id,
            stripe_email=request.user.email,
            amount=amt_in_cents / 100
        )


def delete_subscription(revolv_profile, donation_type):
    if donation_type == 'SOLAR_SEED_FUND':
        subscription = StripeDetails.objects.get(user=revolv_profile, donation_amount__gt=0.0)
    else:
        subscription = StripeDetails.objects.get(user=revolv_profile, amount__gt=0.0)
    subscription = stripe.Subscription.retrieve(subscription.subscription_id)
    customer = subscription.customer
    plan_id = subscription.plan.id
    plan = stripe.Plan.retrieve(plan_id)
    plan.delete()
    subscription.delete()
    if donation_type == 'SOLAR_SEED_FUND':
        StripeDetails.objects.get(user=revolv_profile, donation_amount__gt=0.0).delete()
    else:
        StripeDetails.objects.get(user=revolv_profile, amount__gt=0.0).delete()
    return customer


@login_required
def donation_update(request):
    try:
        operation_amt = request.POST.get('operation-amt')
        donation_amt = request.POST.get('donation-amt')
        operation_amt_cents = round(float(operation_amt) * 100)
        donation_amt_cents = round(float(donation_amt) * 100)
    except KeyError:
        logger.exception('stripe_payment called without required POST data')
        return HttpResponseBadRequest('bad POST data')

    revolv_profile = RevolvUserProfile.objects.get(user=request.user)
    stripedetail = StripeDetails.objects.filter(user=revolv_profile)
    if stripedetail:
        try:
            donation_amount = 0
            if float(operation_amt) <= 0:
                try:
                    donation_type = 'OPERATION'
                    delete_subscription(revolv_profile, donation_type)
                except StripeDetails.DoesNotExist:
                    subscription = None
            else:
                try:
                    amount = StripeDetails.objects.get(user=revolv_profile, amount__gt=0.0).amount
                except StripeDetails.DoesNotExist:
                    amount = 0
                if not abs(float(operation_amt) - float(amount)) < 0.00000001:
                    try:
                        donation_type = 'OPERATION'
                        customer = delete_subscription(revolv_profile, donation_type)
                    except StripeDetails.DoesNotExist:
                        user = StripeDetails.objects.get(user=revolv_profile, donation_amount__gt=0.0)
                        subscription = stripe.Subscription.retrieve(user.subscription_id)
                        customer = subscription.customer

                    donation_amount = float(donation_amount) + float(operation_amt)
                    create_subscription(request, donation_type, revolv_profile, operation_amt_cents, customer)

            if float(donation_amt) <= 0:
                try:
                    donation_type = 'SOLAR_SEED_FUND'
                    delete_subscription(revolv_profile, donation_type)
                except StripeDetails.DoesNotExist:
                    subscription = None
            else:
                try:
                    amount = StripeDetails.objects.get(user=revolv_profile, donation_amount__gt=0.0).donation_amount
                except StripeDetails.DoesNotExist:
                    amount = 0
                if not abs(float(donation_amt) - float(amount)) < 0.00000001:
                    try:
                        donation_type = 'SOLAR_SEED_FUND'
                        customer = delete_subscription(revolv_profile, donation_type)
                    except StripeDetails.DoesNotExist:
                        user = StripeDetails.objects.get(user=revolv_profile, amount__gt=0.0)
                        subscription = stripe.Subscription.retrieve(user.subscription_id)
                        customer = subscription.customer

                    donation_type = 'SOLAR_SEED_FUND'
                    donation_amount = float(donation_amount) + float(donation_amt)
                    create_subscription(request, donation_type, revolv_profile, donation_amt_cents, customer)


        except KeyError:
            logger.exception('stripe_payment called without required POST data')
            return HttpResponseBadRequest('bad POST data')
        messages.success(request, 'Donation details successfully updated')
        return HttpResponseRedirect('/account_settings/')

    else:
        token = request.POST['stripeToken']
        email = request.POST['stripeEmail']
        try:
            donation_amount = 0
            customer = stripe.Customer.create(
                email=email,
                description="Donation for RE-volv Operations",
                source=token  # obtained with Stripe.js
            )

            if float(operation_amt) > 0.0:
                donation_type = 'OPERATION'
                donation_amount = float(donation_amount) + float(operation_amt)
                create_subscription(request, donation_type, revolv_profile, operation_amt_cents, customer["id"])

            if float(donation_amt) > 0.0:
                donation_type = 'SOLAR_SEED_FUND'
                donation_amount = float(donation_amount) + float(donation_amt)
                create_subscription(request, donation_type, revolv_profile, donation_amt_cents, customer["id"])


        except stripe.error.CardError as e:
            body = e.json_body
            # error_msg = body['error']['message']
            messages.error(request, 'Payment error')
            return redirect('home')
        except stripe.error.APIConnectionError as e:
            body = e.json_body
            # error_msg = body['error']['message']
            messages.error(request, 'Internet connection error. Please check your internet connection.')
            return redirect('home')
        except Exception:
            error_msg = "Payment error. RE-volv has been notified."
            logger.exception(error_msg)

            messages.error(request, 'Payment error. RE-volv has been notified.')
            return redirect('home')
        return HttpResponse(json.dumps({'status': 'donation_success', 'amount': donation_amount}),
                            content_type="application/json")


class DonationSourceTrackingView(UserDataMixin, TemplateView):
    """
    Display donation source tracking report

    Accessed through /donationsourcetracking/
    """
    model = Payment
    template_name = 'base/partials/donation_source_tracking.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(DonationSourceTrackingView, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(DonationSourceTrackingView, self).get_context_data(**kwargs)
        sources = ReferralSourceTrack.objects.values('source', 'medium', 'payment_id__project__title')\
            .annotate(amount=Sum('payment__amount')).annotate(count=Count('payment__id'))\
            .order_by('source', 'payment_id__project__title')
        context["sources"] = sources
        return context


class MonthlyRepaymentReport(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    model = Payment
    template_name = 'base/partials/monthly_repayment_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(MonthlyRepaymentReport, self).dispatch(request, *args, **kwargs)
    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(MonthlyRepaymentReport, self).get_context_data(**kwargs)
        return context


def monthly_repayment_table(request):
    import calendar

    draw = request.GET.get('draw')
    length = request.GET.get('length')
    order = request.GET.get('order[0][dir]')
    start = request.GET.get('start')
    search = request.GET.get('search[value]')
    currentSortByCol = request.GET.get('order[0][column]')

    fields = ['created_at', 'project', 'amount']

    order_by = {'asc': '-', 'desc': ''}

    column_order = order_by[order] + fields[int(currentSortByCol)]

    monthly_repayment_list = AdminRepayment.objects.filter(amount__gt=0.00).order_by(column_order)
    admin_repayment_list = []

    admin_repayment_list.append(AdminRepayment.objects.all())

    if search.strip():
        monthly_repayment_list = monthly_repayment_list.filter(Q(project__title__icontains=search) |
                                                               Q(amount__icontains=search) |
                                                               Q(created_at__icontains=search))
    if int(length) == -1:
        length = monthly_repayment_list.count()
    payments = []

    for payment in monthly_repayment_list[int(start):][:int(length)]:
        month = calendar.month_name[payment.created_at.month]
        payment_details={}
        payment_details['month'] = str(month)
        payment_details['year'] = int(payment.created_at.year)
        payment_details['project_name'] = payment.project.title
        payment_details['repayment_amount'] = float(round(payment.amount, 2))
        payments.append(payment_details)

    json_response = {"draw": draw, "recordsTotal": AdminRepayment.objects.all().count(),
                     "recordsFiltered": monthly_repayment_list.count(), "data": payments}

    return HttpResponse(json.dumps(json_response), content_type='application/json')


class UserListView(UserDataMixin, TemplateView):
    """
    Display user's details report

    Accessed through /UserListView/
    """
    model = RevolvUserProfile
    template_name = 'base/partials/user_details.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        if not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        return super(UserListView, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(UserListView, self).get_context_data(**kwargs)
        user_list = RevolvUserProfile.objects.all()
        context["users"] = user_list.order_by("-user__date_joined")
        return context


class DonationHistory(UserDataMixin, TemplateView):
    """
    Display user's donation history

    Accessed through /DonationHistory/
    """
    model = Payment
    template_name = 'base/partials/donation_history.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous():
            return HttpResponseRedirect(reverse("login"))
        return super(DonationHistory, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        context = super(DonationHistory, self).get_context_data(**kwargs)
        donation_list = Payment.objects.filter(user__user__username=self.user_profile)
        context["donation_history"] = donation_list
        return context


class SolarWeekEventReportView(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    template_name = 'base/partials/solar_ed_week_event_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous() or not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        else:
            return super(SolarWeekEventReportView, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        event_records = HostEvent.objects.all()
        context = super(SolarWeekEventReportView, self).get_context_data(**kwargs)
        context['total'] = event_records.count()
        return context


def solar_ed_event_data_table(request):
    draw = request.GET.get('draw')
    length = request.GET.get('length')
    start = request.GET.get('start')
    search = request.GET.get('search[value]')
    order = request.GET.get('order[0][dir]')
    current_sort_by_col = request.GET.get('order[0][column]')
    order_by = {'desc': '-', 'asc': ''}

    fields = ['name', 'email', 'title', 'date',
              'address', 'city', 'amount', 'state', 'zip_code']

    column_order = order_by[order] + fields[int(current_sort_by_col)]

    event_list = HostEvent.objects.all().order_by(column_order)
    if int(length) == -1:
        length = event_list.count()
    if search:
        event_list = event_list.filter(Q(name__icontains=search) |
                                       Q(email__icontains=search) |
                                       Q(title__icontains=search) |
                                       Q(date__icontains=search) |
                                       Q(address__icontains=search) |
                                       Q(city__icontains=search) |
                                       Q(state__icontains=search) |
                                       Q(zip_code__icontains=search) |
                                       Q(detail__icontains=search))

    events = []
    for event in event_list[int(start):][:int(length)]:

        event_details = dict()
        event_details['name'] = event.name
        event_details['email'] = event.email
        event_details['title'] = event.title
        event_details['date'] = str(event.date)
        event_details['address'] = event.address
        event_details['city'] = event.city
        event_details['state'] = event.state
        event_details['zip_code'] = event.zip_code
        event_details['detail'] = event.detail
        events.append(event_details)
    json_response = {"draw": draw, "recordsTotal": event_list.count(),
                     "recordsFiltered": event_list.count(), "all-data": events}

    return HttpResponse(json.dumps(json_response), content_type='application/json')


class SolarWeekPartnerReportView(UserDataMixin, TemplateView):
    """
    The project view. Displays project details and allows for editing.

    Accessed through /project/{project_id}
    """
    template_name = 'base/partials/solar_ed_week_partners_report.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous() or not request.user.revolvuserprofile.is_administrator():
            return HttpResponseRedirect(reverse("dashboard"))
        else:
            return super(SolarWeekPartnerReportView, self).dispatch(request, *args, **kwargs)

    # pass in Project Categories and Maps API key
    def get_context_data(self, **kwargs):
        event_records = HostEvent.objects.all()
        context = super(SolarWeekPartnerReportView, self).get_context_data(**kwargs)
        context['total'] = event_records.count()
        return context


def solar_ed_partner_data_table(request):
    draw = request.GET.get('draw')
    length = request.GET.get('length')
    start = request.GET.get('start')
    search = request.GET.get('search[value]')
    order = request.GET.get('order[0][dir]')
    current_sort_by_col = request.GET.get('order[0][column]')
    order_by = {'desc': '-', 'asc': ''}

    fields = ['name', 'email', 'organization', 'promoting_way']

    column_order = order_by[order] + fields[int(current_sort_by_col)]

    partner_list = BecomePartner.objects.all().order_by(column_order)
    if int(length) == -1:
        length = partner_list.count()
    if search:
        partner_list = partner_list.filter(Q(name__icontains=search) |
                                           Q(email__icontains=search) |
                                           Q(organization__icontains=search) |
                                           Q(promoting_way__icontains=search))

    partners = []
    for partner in partner_list[int(start):][:int(length)]:

        partner_details = dict()
        partner_details['name'] = partner.name
        partner_details['email'] = partner.email
        partner_details['organization'] = partner.organization
        partner_details['promoting_way'] = partner.promoting_way
        partners.append(partner_details)
    json_response = {"draw": draw, "recordsTotal": partner_list.count(),
                     "recordsFiltered": partner_list.count(), "all-data": partners}

    return HttpResponse(json.dumps(json_response), content_type='application/json')